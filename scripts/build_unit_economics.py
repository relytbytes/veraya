from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = "0000FF"; BLACK = "000000"; GREEN = "008000"; WHITE = "FFFFFF"
NAVY = "0C1A1E"; TEAL = "21A090"
YEL = "FFF3B0"; HDR = "0C1A1E"; SUB = "E8F2F0"

f_title = Font(name="Arial", size=15, bold=True, color=WHITE)
f_sub   = Font(name="Arial", size=9, italic=True, color="555555")
f_sect  = Font(name="Arial", size=11, bold=True, color=WHITE)
f_lbl   = Font(name="Arial", size=10, color=BLACK)
f_in    = Font(name="Arial", size=10, bold=True, color=BLUE)
f_form  = Font(name="Arial", size=10, color=BLACK)
f_link  = Font(name="Arial", size=10, color=GREEN)
f_colh  = Font(name="Arial", size=10, bold=True, color=BLACK)
f_note  = Font(name="Arial", size=9, color="555555")
f_bold  = Font(name="Arial", size=10, bold=True, color=BLACK)

fill_title = PatternFill("solid", fgColor=NAVY)
fill_sect  = PatternFill("solid", fgColor=TEAL)
fill_yel   = PatternFill("solid", fgColor=YEL)
fill_sub   = PatternFill("solid", fgColor=SUB)
thin = Side(style="thin", color="CCCCCC")
border = Border(bottom=thin)

PCT = "0.00%"; CUR = '$#,##0;($#,##0);"-"'; CUR2 = '$#,##0.00'; CNT = "#,##0"
R = Alignment(horizontal="right"); L = Alignment(horizontal="left"); C = Alignment(horizontal="center")

wb = Workbook()

# ───────────────────────── Sheet 1: Model ─────────────────────────
ws = wb.active; ws.title = "Model"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 46
for col in "BCD": ws.column_dimensions[col].width = 17
ws.column_dimensions["E"].width = 3
ws.column_dimensions["F"].width = 40

def cell(addr, val, font=f_form, fmt=None, fill=None, align=None):
    c = ws[addr]; c.value = val; c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = align
    return c

def sect(row, text):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws[f"A{row}"]; c.value = text; c.font = f_sect; c.fill = fill_sect; c.alignment = L

ws.merge_cells("A1:D1"); ws["A1"].value = "Veraya — Payments & SaaS Unit Economics"
ws["A1"].font = f_title; ws["A1"].fill = fill_title; ws["A1"].alignment = L
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:D2"); ws["A2"].value = "Change the BLUE inputs; everything recalculates. YELLOW cells are the key assumptions to revisit. Defaults model a solid independent (~$1.5M/yr)."
ws["A2"].font = f_sub; ws["A2"].alignment = L

# INPUTS
sect(4, "INPUTS — the operator & your pricing")
cell("A5", "Restaurants (live instances)", f_lbl); cell("B5", 10, f_in, CNT, align=R)
cell("A6", "Avg card volume / restaurant ($/mo)", f_lbl); cell("B6", 120000, f_in, CUR, align=R)
cell("A7", "Avg ticket / check ($)", f_lbl); cell("B7", 55, f_in, CUR, align=R)
cell("A8", "→ Transactions / restaurant / mo", f_lbl); cell("B8", "=B6/B7", f_form, CNT, align=R)
cell("A10", "SaaS price — base, Models A & B ($/mo)", f_lbl); cell("B10", 349, f_in, CUR, align=R)
cell("A11", "SaaS price — membership, Model C ($/mo)", f_lbl); cell("B11", 549, f_in, CUR, align=R)
cell("A12", "Infra cost / restaurant ($/mo)", f_lbl); cell("B12", 40, f_in, CUR, align=R)

# COST BASIS
sect(14, "YOUR PROCESSING COST  (what the rails/payfac charge YOU, incl. interchange)")
cell("A15", "Processing cost — rate (% of volume)", f_lbl); cell("B15", 0.021, f_in, PCT, fill_yel, align=R)
cell("A16", "Processing cost — per transaction ($)", f_lbl); cell("B16", 0.08, f_in, CUR2, fill_yel, align=R)
cell("A17", "→ Your processing cost / restaurant ($/mo)", f_bold); cell("B17", "=B15*B6+B16*B8", f_form, CUR, align=R)
cell("F15", "Interchange-plus basis (~2.1%). On vanilla Stripe flat (~2.7%+$0.05) you can't beat Square — the savings story REQUIRES volume interchange-plus.", f_note, align=L)

# MERCHANT PRICING MODELS
sect(19, "MERCHANT PRICING MODELS — what YOU charge the restaurant")
cell("A20", "", f_lbl); cell("B20", "Rate (%)", f_colh, align=R); cell("C20", "Per txn ($)", f_colh, align=R)
cell("A21", "Model A — Flat % (Square / Toast style)", f_lbl); cell("B21", 0.026, f_in, PCT, fill_yel, align=R); cell("C21", 0.15, f_in, CUR2, fill_yel, align=R)
cell("A22", "Model B — Interchange-plus markup (thin)", f_lbl); cell("B22", 0.0015, f_in, PCT, fill_yel, align=R); cell("C22", 0.05, f_in, CUR2, fill_yel, align=R)
cell("A23", "Model C — Membership (0% markup)", f_lbl); cell("B23", 0.0, f_in, PCT, fill_yel, align=R); cell("C23", 0.0, f_in, CUR2, fill_yel, align=R)
cell("F21", "A: you behave like Square — most revenue, least on-brand.", f_note, align=L)
cell("F22", "B: cost + thin markup — fair, transparent, still beats Square.", f_note, align=L)
cell("F23", "C: process at cost, earn on a higher SaaS tier — 'we don't tax sales.'", f_note, align=L)

# COMPARISON TABLE
sect(25, "PER-RESTAURANT ECONOMICS  (monthly)")
cell("A26", "Metric", f_colh, fill=fill_sub); cell("B26", "A: Flat %", f_colh, fill=fill_sub, align=R); cell("C26", "B: Intchg-plus", f_colh, fill=fill_sub, align=R); cell("D26", "C: Membership", f_colh, fill=fill_sub, align=R)
cell("A27", "Merchant pays — processing ($/mo)", f_lbl)
cell("B27", "=B21*$B$6+C21*$B$8", f_form, CUR, align=R)
cell("C27", "=$B$17+(B22*$B$6+C22*$B$8)", f_form, CUR, align=R)
cell("D27", "=$B$17+(B23*$B$6+C23*$B$8)", f_form, CUR, align=R)
cell("A28", "Merchant effective rate (%)", f_lbl)
for col in "BCD": cell(f"{col}28", f"={col}27/$B$6", f_form, PCT, align=R)
cell("A29", "Your PAYMENTS revenue ($/mo)", f_lbl)
for col in "BCD": cell(f"{col}29", f"={col}27-$B$17", f_form, CUR, align=R)
cell("A30", "Your SaaS revenue ($/mo)", f_lbl)
cell("B30", "=$B$10", f_form, CUR, align=R); cell("C30", "=$B$10", f_form, CUR, align=R); cell("D30", "=$B$11", f_form, CUR, align=R)
cell("A31", "Your TOTAL revenue ($/mo)", f_bold)
for col in "BCD": cell(f"{col}31", f"={col}29+{col}30", f_bold, CUR, align=R)
cell("A32", "Your contribution after infra ($/mo)", f_lbl)
for col in "BCD": cell(f"{col}32", f"={col}31-$B$12", f_form, CUR, align=R)

sect(34, "PER-RESTAURANT  (annual)  &  FLEET")
cell("A35", "Your total revenue / restaurant ($/yr)", f_bold)
for col in "BCD": cell(f"{col}35", f"={col}31*12", f_bold, CUR, align=R)
cell("A36", "Fleet total revenue — × restaurants ($/yr)", f_bold)
for col in "BCD": cell(f"{col}36", f"={col}35*$B$5", f_bold, CUR, align=R)
cell("A37", "Fleet contribution after infra ($/yr)", f_lbl)
for col in "BCD": cell(f"{col}37", f"={col}32*12*$B$5", f_form, CUR, align=R)

sect(39, "MERCHANT VIEW  (the sales pitch)")
cell("A40", "Merchant all-in to Veraya — proc.+SaaS ($/mo)", f_lbl)
for col in "BCD": cell(f"{col}40", f"={col}27+{col}30", f_form, CUR, align=R)
cell("A41", "Merchant processing savings vs Square ($/mo)", f_lbl)
for col in "BCD": cell(f"{col}41", f"=$B$27-{col}27", f_form, CUR, align=R)
cell("A42", "Merchant processing savings vs Square ($/yr)", f_lbl)
for col in "BCD": cell(f"{col}42", f"={col}41*12", f_form, CUR, align=R)

sect(44, "PER AVERAGE CHECK  (at avg ticket — the whiteboard example)")
cell("A45", "Merchant pays per check ($)", f_lbl)
cell("B45", "=B21*$B$7+C21", f_form, CUR2, align=R)
cell("C45", "=($B$15*$B$7+$B$16)+(B22*$B$7+C22)", f_form, CUR2, align=R)
cell("D45", "=($B$15*$B$7+$B$16)+(B23*$B$7+C23)", f_form, CUR2, align=R)
cell("A46", "Effective rate per check (%)", f_lbl)
for col in "BCD": cell(f"{col}46", f"={col}45/$B$7", f_form, PCT, align=R)
cell("A47", "Saved per check vs Square ($)", f_lbl)
for col in "BCD": cell(f"{col}47", f"=$B$45-{col}45", f_form, CUR2, align=R)

# ───────────────────────── Sheet 2: Fleet & Scaling ─────────────────────────
fs = wb.create_sheet("Fleet & Scaling")
fs.sheet_view.showGridLines = False
fs.column_dimensions["A"].width = 26
for col in "BCDE": fs.column_dimensions[col].width = 18

def fcell(addr, val, font=f_form, fmt=None, fill=None, align=None):
    c = fs[addr]; c.value = val; c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = align
    return c

fs.merge_cells("A1:E1"); fs["A1"].value = "Fleet & Scaling Ramp"
fs["A1"].font = f_title; fs["A1"].fill = fill_title; fs["A1"].alignment = L
fs.row_dimensions[1].height = 24

fs.merge_cells("A3:E3"); fs["A3"].value = "RAMP — recommended model: B (interchange-plus). Per-unit economics pulled from the Model sheet."
fs["A3"].font = f_sect; fs["A3"].fill = fill_sect; fs["A3"].alignment = L
hdrs = ["Restaurants", "SaaS rev ($/yr)", "Payments rev ($/yr)", "Total rev ($/yr)", "Contribution ($/yr)"]
for i, h in enumerate(hdrs):
    fcell(f"{get_column_letter(1+i)}4", h, f_colh, fill=fill_sub, align=(L if i == 0 else R))
counts = [1, 3, 5, 10, 15, 20]
r = 5
for n in counts:
    fcell(f"A{r}", n, f_in, CNT, align=L)
    fcell(f"B{r}", f"=A{r}*Model!$C$30*12", f_link, CUR, align=R)
    fcell(f"C{r}", f"=A{r}*Model!$C$29*12", f_link, CUR, align=R)
    fcell(f"D{r}", f"=A{r}*Model!$C$35", f_link, CUR, align=R)
    fcell(f"E{r}", f"=A{r}*Model!$C$32*12", f_link, CUR, align=R)
    r += 1
fcell(f"A{r+1}", "Solo + AI operating ceiling ≈ 15–20 live instances before support/ops needs a hire.", f_note, align=L)
fs.merge_cells(f"A{r+1}:E{r+1}")

fs.merge_cells("A14:E14"); fs["A14"].value = "ALL THREE MODELS at current inputs (fleet, $/yr)"
fs["A14"].font = f_sect; fs["A14"].fill = fill_sect; fs["A14"].alignment = L
fcell("A15", "Model", f_colh, fill=fill_sub); fcell("B15", "A: Flat %", f_colh, fill=fill_sub, align=R); fcell("C15", "B: Intchg-plus", f_colh, fill=fill_sub, align=R); fcell("D15", "C: Membership", f_colh, fill=fill_sub, align=R)
fcell("A16", "Fleet total revenue ($/yr)", f_lbl)
for col in "BCD": fcell(f"{col}16", f"=Model!{col}36", f_link, CUR, align=R)
fcell("A17", "Fleet contribution ($/yr)", f_lbl)
for col in "BCD": fcell(f"{col}17", f"=Model!{col}37", f_link, CUR, align=R)
fcell("A18", "Of total: from payments (%)", f_lbl)
for col in "BCD": fcell(f"{col}18", f"=IF(Model!{col}31=0,0,Model!{col}29/Model!{col}31)", f_link, PCT, align=R)
fcell("A20", "Read: Model A earns the most but taxes sales like Square; C earns least but is purest 'we charge for software'. B balances. Watch row 18 — a % markup on six-figure volume can quietly become your biggest line (that's why Toast leans on payments).", f_note, align=L)
fs.merge_cells("A20:E22")
fs["A20"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ───────────────────────── Sheet 3: Assumptions & Notes ─────────────────────────
nt = wb.create_sheet("Assumptions & Notes")
nt.sheet_view.showGridLines = False
nt.column_dimensions["A"].width = 110
nt.merge_cells("A1:A1"); nt["A1"].value = "Assumptions, Rate References & Strategy Notes"
nt["A1"].font = f_title; nt["A1"].fill = fill_title; nt["A1"].alignment = L
nt.row_dimensions[1].height = 24
notes = [
 ("", ""),
 ("THE COST STACK (every card txn)", "sect"),
 ("• Interchange (issuing bank): ~1.5–2.2% + ~$0.10 card-present. Non-negotiable.", ""),
 ("• Network assessments (Visa/MC): ~0.13–0.14%. Non-negotiable.", ""),
 ("• Processor / payfac markup: the only negotiable layer. Interchange (~2.0–2.3% all-in) is the floor nobody beats.", ""),
 ("", ""),
 ("MARKET REFERENCE RATES (merchant-facing, card-present)", "sect"),
 ("• Square: 2.6% + $0.15 in-person; 2.9% + $0.30 online.", ""),
 ("• Stripe: Terminal ~2.7% + $0.05 card-present; 2.9% + $0.30 online.", ""),
 ("• Toast: custom; ~2.49–2.99% + $0.15; offers interchange-plus to larger merchants.", ""),
 ("Source: public published rates, 2024–2025; rates move — treat as order-of-magnitude.", "note"),
 ("", ""),
 ("YOUR COST BASIS (Model sheet, yellow cells)", "sect"),
 ("• Default 2.10% + $0.08 = interchange-plus via a payfac at modest volume.", ""),
 ("• IMPORTANT: on vanilla Stripe flat (~2.7% + $0.05) you cannot price below Square AND make margin.", ""),
 ("  The 'we save you money on processing' story only works once you're on true interchange-plus pricing.", ""),
 ("", ""),
 ("THE STRATEGIC TRADE-OFF (why three models)", "sect"),
 ("• Model A (flat %): you behave like Square/Toast — highest revenue, but you 'tax' the operator's sales. Off-brand.", ""),
 ("• Model B (interchange-plus, thin markup): fair + transparent, still beats Square, modest payments margin.", ""),
 ("• Model C (membership): process at cost, earn on a higher SaaS tier. Best merchant deal, cleanest narrative,", ""),
 ("  lowest payments revenue. This is the purest expression of 'we charge for the brain, not your sales.'", ""),
 ("• Watch 'Fleet & Scaling' row 18: a 'thin' % markup on six-figure monthly volume can exceed your SaaS line.", ""),
 ("  That's exactly why Toast leans on payments — decide deliberately whether you want to.", ""),
 ("", ""),
 ("RECOMMENDATION", "sect"),
 ("Lead with Model B or C. Price payments to be non-objectionable (fair/transparent), not to be the hero.", ""),
 ("Veraya's margin should come from the platform + Vera. Use the $150-check math (Model sheet, row 45) in the pitch.", ""),
 ("", ""),
 ("CAVEATS", "sect"),
 ("• Payments margin carries chargeback/settlement/PCI-light obligations — riding Stripe/payfac keeps most of that off you.", ""),
 ("• You will always have a processor relationship (the one unavoidable outside dependency); it's paid as a % of txns, not upfront.", ""),
 ("• Excludes: hardware, chargeback losses, support cost per account, taxes. This models revenue/contribution, not full P&L.", ""),
]
r = 2
for text, kind in notes:
    c = nt[f"A{r}"]; c.value = text
    if kind == "sect": c.font = f_sect; c.fill = fill_sect
    elif kind == "note": c.font = f_note
    else: c.font = f_lbl
    c.alignment = L
    r += 1

wb.save("/Users/ty/restaurant-ops/docs/business/Veraya_Unit_Economics.xlsx")
print("saved")
